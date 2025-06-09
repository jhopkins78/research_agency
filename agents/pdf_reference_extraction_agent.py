"""
PDF Reference Extraction Agent (PREA)
A specialized agent for extracting academic references from PDF documents
and storing them in structured formats (files and spreadsheets).

This agent integrates with the Academic Research Agent System (ARAS) to provide
comprehensive PDF document analysis and reference extraction capabilities.

"""

import re
import json
import csv
import time
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, asdict
import os
from pathlib import Path

# PDF processing imports
try:
    import PyPDF2
    import pdfplumber
    from pdf2image import convert_from_path
    import pytesseract
    PDF_LIBRARIES_AVAILABLE = True
except ImportError:
    PDF_LIBRARIES_AVAILABLE = False
    print("Warning: PDF processing libraries not available. Install with: pip install PyPDF2 pdfplumber pdf2image pytesseract")

# Spreadsheet generation imports
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import pandas as pd
    SPREADSHEET_LIBRARIES_AVAILABLE = True
except ImportError:
    SPREADSHEET_LIBRARIES_AVAILABLE = False
    print("Warning: Spreadsheet libraries not available. Install with: pip install openpyxl pandas")

@dataclass
class ExtractedReference:
    """Data structure for extracted reference information"""
    reference_number: Optional[int] = None
    full_text: str = ""
    authors: List[str] = None
    title: str = ""
    year: Optional[int] = None
    venue: str = ""  # Journal, conference, or publisher
    volume: str = ""
    issue: str = ""
    pages: str = ""
    doi: str = ""
    url: str = ""
    isbn: str = ""
    reference_type: str = ""  # journal, book, conference, website, etc.
    citation_style: str = ""  # apa, mla, chicago, etc.
    confidence_score: float = 0.0
    extraction_notes: str = ""
    
    def __post_init__(self):
        if self.authors is None:
            self.authors = []

class PDFTextExtractor:
    """
    Handles extraction of text from PDF documents using multiple methods
    """
    
    def __init__(self, config: Dict):
        self.config = config
        self.extraction_methods = ["pdfplumber", "pypdf2", "ocr"]
        
    def extract_text_from_pdf(self, pdf_path: str) -> Dict[str, any]:
        """
        Extract text from PDF using multiple methods and return the best result
        """
        if not PDF_LIBRARIES_AVAILABLE:
            raise ImportError("PDF processing libraries not available")
            
        extraction_results = {}
        
        # Try different extraction methods
        for method in self.extraction_methods:
            try:
                if method == "pdfplumber":
                    result = self._extract_with_pdfplumber(pdf_path)
                elif method == "pypdf2":
                    result = self._extract_with_pypdf2(pdf_path)
                elif method == "ocr":
                    result = self._extract_with_ocr(pdf_path)
                else:
                    continue
                    
                extraction_results[method] = result
                
            except Exception as e:
                print(f"Error with {method}: {str(e)}")
                extraction_results[method] = {
                    "text": "",
                    "pages": [],
                    "success": False,
                    "error": str(e)
                }
        
        # Select the best extraction result
        best_result = self._select_best_extraction(extraction_results)
        
        return {
            "text": best_result.get("text", ""),
            "pages": best_result.get("pages", []),
            "method_used": best_result.get("method", "unknown"),
            "extraction_quality": best_result.get("quality_score", 0.0),
            "all_results": extraction_results
        }
    
    def _extract_with_pdfplumber(self, pdf_path: str) -> Dict[str, any]:
        """Extract text using pdfplumber (best for most PDFs)"""
        text_content = []
        pages_content = []
        
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                page_text = page.extract_text()
                if page_text:
                    text_content.append(page_text)
                    pages_content.append({
                        "page_number": page_num + 1,
                        "text": page_text,
                        "char_count": len(page_text)
                    })
        
        full_text = "\n".join(text_content)
        
        return {
            "text": full_text,
            "pages": pages_content,
            "method": "pdfplumber",
            "success": True,
            "quality_score": self._calculate_text_quality(full_text)
        }
    
    def _extract_with_pypdf2(self, pdf_path: str) -> Dict[str, any]:
        """Extract text using PyPDF2 (fallback method)"""
        text_content = []
        pages_content = []
        
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            for page_num, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text()
                if page_text:
                    text_content.append(page_text)
                    pages_content.append({
                        "page_number": page_num + 1,
                        "text": page_text,
                        "char_count": len(page_text)
                    })
        
        full_text = "\n".join(text_content)
        
        return {
            "text": full_text,
            "pages": pages_content,
            "method": "pypdf2",
            "success": True,
            "quality_score": self._calculate_text_quality(full_text)
        }
    
    def _extract_with_ocr(self, pdf_path: str) -> Dict[str, any]:
        """Extract text using OCR (for scanned PDFs)"""
        try:
            # Convert PDF to images
            images = convert_from_path(pdf_path)
            
            text_content = []
            pages_content = []
            
            for page_num, image in enumerate(images):
                # Use OCR to extract text from image
                page_text = pytesseract.image_to_string(image)
                if page_text.strip():
                    text_content.append(page_text)
                    pages_content.append({
                        "page_number": page_num + 1,
                        "text": page_text,
                        "char_count": len(page_text)
                    })
            
            full_text = "\n".join(text_content)
            
            return {
                "text": full_text,
                "pages": pages_content,
                "method": "ocr",
                "success": True,
                "quality_score": self._calculate_text_quality(full_text) * 0.8  # OCR typically lower quality
            }
            
        except Exception as e:
            return {
                "text": "",
                "pages": [],
                "method": "ocr",
                "success": False,
                "error": str(e),
                "quality_score": 0.0
            }
    
    def _calculate_text_quality(self, text: str) -> float:
        """Calculate a quality score for extracted text"""
        if not text:
            return 0.0
        
        # Basic quality metrics
        char_count = len(text)
        word_count = len(text.split())
        line_count = len(text.split('\n'))
        
        # Check for common academic indicators
        academic_indicators = [
            'abstract', 'introduction', 'methodology', 'results', 'conclusion',
            'references', 'bibliography', 'doi:', 'http://', 'https://',
            'journal', 'conference', 'proceedings', 'volume', 'issue'
        ]
        
        indicator_count = sum(1 for indicator in academic_indicators if indicator.lower() in text.lower())
        
        # Calculate quality score (0.0 to 1.0)
        quality_score = min(1.0, (
            (char_count / 10000) * 0.3 +  # Length factor
            (word_count / 2000) * 0.3 +   # Word density
            (indicator_count / len(academic_indicators)) * 0.4  # Academic content
        ))
        
        return quality_score
    
    def _select_best_extraction(self, extraction_results: Dict[str, Dict]) -> Dict[str, any]:
        """Select the best extraction result based on quality scores"""
        best_result = {"quality_score": 0.0, "text": "", "method": "none"}
        
        for method, result in extraction_results.items():
            if result.get("success", False) and result.get("quality_score", 0.0) > best_result["quality_score"]:
                best_result = result
        
        return best_result

class ReferenceParser:
    """
    Parses extracted text to identify and structure academic references
    """
    
    def __init__(self, config: Dict):
        self.config = config
        self.citation_patterns = self._initialize_citation_patterns()
        
    def _initialize_citation_patterns(self) -> Dict[str, List[str]]:
        """Initialize regex patterns for different citation styles"""
        return {
            "apa": [
                # Author, A. A. (Year). Title. Journal, Volume(Issue), pages.
                r'([A-Z][a-z]+(?:,\s*[A-Z]\.(?:\s*[A-Z]\.)?)*)\s*\((\d{4})\)\.\s*([^.]+)\.\s*([^,]+)(?:,\s*(\d+)(?:\((\d+)\))?)(?:,\s*([\d-]+))?',
                # Author, A. A., & Author, B. B. (Year). Book title. Publisher.
                r'([A-Z][a-z]+(?:,\s*[A-Z]\.(?:\s*[A-Z]\.)?)*(?:,?\s*&\s*[A-Z][a-z]+(?:,\s*[A-Z]\.(?:\s*[A-Z]\.)?)*)*)\s*\((\d{4})\)\.\s*([^.]+)\.\s*([^.]+)\.'
            ],
            "mla": [
                # Author, First. "Title." Journal, vol. #, no. #, Year, pp. #-#.
                r'([A-Z][a-z]+,\s*[A-Z][a-z]+)\.\s*"([^"]+)"\.\s*([^,]+),\s*vol\.\s*(\d+)(?:,\s*no\.\s*(\d+))?,\s*(\d{4}),\s*pp\.\s*([\d-]+)',
                # Author, First. Book Title. Publisher, Year.
                r'([A-Z][a-z]+,\s*[A-Z][a-z]+)\.\s*([^.]+)\.\s*([^,]+),\s*(\d{4})\.'
            ],
            "chicago": [
                # Author, First Last. "Title." Journal Volume, no. Issue (Year): pages.
                r'([A-Z][a-z]+,\s*[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\.\s*"([^"]+)"\.\s*([^0-9]+)\s*(\d+)(?:,\s*no\.\s*(\d+))?\s*\((\d{4})\):\s*([\d-]+)',
                # Author, First Last. Book Title. Place: Publisher, Year.
                r'([A-Z][a-z]+,\s*[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\.\s*([^.]+)\.\s*[^:]+:\s*([^,]+),\s*(\d{4})\.'
            ],
            "ieee": [
                # [1] A. Author, "Title," Journal, vol. #, no. #, pp. #-#, Year.
                r'\[(\d+)\]\s*([A-Z]\.\s*[A-Z][a-z]+(?:,\s*[A-Z]\.\s*[A-Z][a-z]+)*),\s*"([^"]+)",\s*([^,]+),\s*vol\.\s*(\d+)(?:,\s*no\.\s*(\d+))?,\s*pp\.\s*([\d-]+),\s*(\d{4})',
                # [1] A. Author, Book Title. Publisher, Year.
                r'\[(\d+)\]\s*([A-Z]\.\s*[A-Z][a-z]+(?:,\s*[A-Z]\.\s*[A-Z][a-z]+)*),\s*([^.]+)\.\s*([^,]+),\s*(\d{4})\.'
            ]
        }
    
    def extract_references_from_text(self, text: str) -> List[ExtractedReference]:
        """
        Extract references from the full text of a document
        """
        references = []
        
        # First, try to find the references section
        references_section = self._find_references_section(text)
        
        if references_section:
            # Parse references from the dedicated section
            references.extend(self._parse_references_section(references_section))
        
        # Also look for in-text citations and numbered references
        references.extend(self._find_numbered_references(text))
        
        # Remove duplicates and improve quality
        references = self._deduplicate_references(references)
        references = self._enhance_reference_quality(references)
        
        return references
    
    def _find_references_section(self, text: str) -> Optional[str]:
        """
        Find and extract the references/bibliography section from the text
        """
        # Common section headers
        section_patterns = [
            r'(?i)\n\s*references\s*\n',
            r'(?i)\n\s*bibliography\s*\n',
            r'(?i)\n\s*works\s+cited\s*\n',
            r'(?i)\n\s*literature\s+cited\s*\n',
            r'(?i)\n\s*citations\s*\n'
        ]
        
        for pattern in section_patterns:
            match = re.search(pattern, text)
            if match:
                # Extract text from the match to the end or next major section
                start_pos = match.end()
                
                # Look for end of references (next major section)
                end_patterns = [
                    r'(?i)\n\s*appendix',
                    r'(?i)\n\s*acknowledgments?',
                    r'(?i)\n\s*author\s+information',
                    r'(?i)\n\s*about\s+the\s+authors?'
                ]
                
                end_pos = len(text)
                for end_pattern in end_patterns:
                    end_match = re.search(end_pattern, text[start_pos:])
                    if end_match:
                        end_pos = start_pos + end_match.start()
                        break
                
                return text[start_pos:end_pos].strip()
        
        return None
    
    def _parse_references_section(self, references_text: str) -> List[ExtractedReference]:
        """
        Parse individual references from the references section
        """
        references = []
        
        # Split into individual references (usually separated by newlines or numbers)
        reference_lines = self._split_into_references(references_text)
        
        for i, ref_text in enumerate(reference_lines):
            if len(ref_text.strip()) < 20:  # Skip very short lines
                continue
                
            extracted_ref = self._parse_single_reference(ref_text, i + 1)
            if extracted_ref:
                references.append(extracted_ref)
        
        return references
    
    def _split_into_references(self, text: str) -> List[str]:
        """
        Split references text into individual reference entries
        """
        # Try different splitting strategies
        
        # Strategy 1: Split by numbered references [1], [2], etc.
        numbered_pattern = r'\[(\d+)\]'
        if re.search(numbered_pattern, text):
            parts = re.split(numbered_pattern, text)
            references = []
            for i in range(1, len(parts), 2):  # Skip the first empty part and take every other
                if i + 1 < len(parts):
                    ref_number = parts[i]
                    ref_text = parts[i + 1].strip()
                    if ref_text:
                        references.append(f"[{ref_number}] {ref_text}")
            return references
        
        # Strategy 2: Split by line breaks and merge multi-line references
        lines = text.split('\n')
        references = []
        current_ref = ""
        
        for line in lines:
            line = line.strip()
            if not line:
                if current_ref:
                    references.append(current_ref)
                    current_ref = ""
                continue
            
            # Check if this line starts a new reference
            if self._is_reference_start(line):
                if current_ref:
                    references.append(current_ref)
                current_ref = line
            else:
                # Continuation of previous reference
                if current_ref:
                    current_ref += " " + line
                else:
                    current_ref = line
        
        # Add the last reference
        if current_ref:
            references.append(current_ref)
        
        return references
    
    def _is_reference_start(self, line: str) -> bool:
        """
        Determine if a line starts a new reference
        """
        # Check for common reference starting patterns
        patterns = [
            r'^\[\d+\]',  # [1], [2], etc.
            r'^\d+\.',    # 1., 2., etc.
            r'^[A-Z][a-z]+,\s*[A-Z]',  # Author, A.
            r'^[A-Z][a-z]+,\s*[A-Z][a-z]+'  # Author, First
        ]
        
        for pattern in patterns:
            if re.match(pattern, line):
                return True
        
        return False
    
    def _parse_single_reference(self, ref_text: str, ref_number: int) -> Optional[ExtractedReference]:
        """
        Parse a single reference text into structured data
        """
        ref_text = ref_text.strip()
        
        # Try different citation style patterns
        for style, patterns in self.citation_patterns.items():
            for pattern in patterns:
                match = re.search(pattern, ref_text)
                if match:
                    return self._create_reference_from_match(match, ref_text, ref_number, style)
        
        # If no pattern matches, create a basic reference
        return ExtractedReference(
            reference_number=ref_number,
            full_text=ref_text,
            confidence_score=0.3,
            extraction_notes="Pattern matching failed, basic extraction only"
        )
    
    def _create_reference_from_match(self, match, ref_text: str, ref_number: int, style: str) -> ExtractedReference:
        """
        Create an ExtractedReference object from a regex match
        """
        groups = match.groups()
        
        # Basic extraction (varies by citation style)
        reference = ExtractedReference(
            reference_number=ref_number,
            full_text=ref_text,
            citation_style=style,
            confidence_score=0.8
        )
        
        # Extract common fields based on the pattern
        if style == "apa" and len(groups) >= 3:
            reference.authors = [groups[0].strip()]
            reference.year = int(groups[1]) if groups[1].isdigit() else None
            reference.title = groups[2].strip()
            if len(groups) > 3:
                reference.venue = groups[3].strip()
        
        elif style == "ieee" and len(groups) >= 4:
            if groups[0].isdigit():  # Numbered reference
                reference.reference_number = int(groups[0])
                reference.authors = [groups[1].strip()]
                reference.title = groups[2].strip()
                reference.venue = groups[3].strip()
        
        # Extract additional metadata
        reference = self._extract_additional_metadata(reference, ref_text)
        
        return reference
    
    def _extract_additional_metadata(self, reference: ExtractedReference, ref_text: str) -> ExtractedReference:
        """
        Extract additional metadata like DOI, URL, ISBN, etc.
        """
        # DOI extraction
        doi_pattern = r'(?:doi:|DOI:)\s*(10\.\d+/[^\s]+)'
        doi_match = re.search(doi_pattern, ref_text, re.IGNORECASE)
        if doi_match:
            reference.doi = doi_match.group(1)
        
        # URL extraction
        url_pattern = r'(https?://[^\s]+)'
        url_match = re.search(url_pattern, ref_text)
        if url_match:
            reference.url = url_match.group(1)
        
        # ISBN extraction
        isbn_pattern = r'(?:ISBN:?\s*)((?:\d{3}-)?\d{1,5}-\d{1,7}-\d{1,7}-[\dX])'
        isbn_match = re.search(isbn_pattern, ref_text, re.IGNORECASE)
        if isbn_match:
            reference.isbn = isbn_match.group(1)
        
        # Volume and issue extraction
        volume_pattern = r'(?:vol\.?\s*|volume\s*)(\d+)'
        volume_match = re.search(volume_pattern, ref_text, re.IGNORECASE)
        if volume_match:
            reference.volume = volume_match.group(1)
        
        issue_pattern = r'(?:no\.?\s*|issue\s*|number\s*)(\d+)'
        issue_match = re.search(issue_pattern, ref_text, re.IGNORECASE)
        if issue_match:
            reference.issue = issue_match.group(1)
        
        # Pages extraction
        pages_pattern = r'(?:pp?\.?\s*|pages?\s*)([\d-]+)'
        pages_match = re.search(pages_pattern, ref_text, re.IGNORECASE)
        if pages_match:
            reference.pages = pages_match.group(1)
        
        # Determine reference type
        reference.reference_type = self._determine_reference_type(ref_text)
        
        return reference
    
    def _determine_reference_type(self, ref_text: str) -> str:
        """
        Determine the type of reference (journal, book, conference, etc.)
        """
        ref_lower = ref_text.lower()
        
        if any(keyword in ref_lower for keyword in ['journal', 'vol.', 'volume', 'issue']):
            return "journal"
        elif any(keyword in ref_lower for keyword in ['proceedings', 'conference', 'symposium']):
            return "conference"
        elif any(keyword in ref_lower for keyword in ['book', 'publisher', 'press']):
            return "book"
        elif any(keyword in ref_lower for keyword in ['http://', 'https://', 'www.']):
            return "website"
        elif 'thesis' in ref_lower or 'dissertation' in ref_lower:
            return "thesis"
        else:
            return "unknown"
    
    def _find_numbered_references(self, text: str) -> List[ExtractedReference]:
        """
        Find numbered references throughout the document (like [1], [2], etc.)
        """
        references = []
        
        # Pattern for numbered references
        pattern = r'\[(\d+)\]\s*([^[\n]+(?:\n[^[\n]+)*?)(?=\[\d+\]|\n\s*\n|$)'
        
        matches = re.finditer(pattern, text, re.MULTILINE)
        
        for match in matches:
            ref_number = int(match.group(1))
            ref_text = match.group(2).strip()
            
            if len(ref_text) > 20:  # Only process substantial references
                extracted_ref = self._parse_single_reference(ref_text, ref_number)
                if extracted_ref:
                    references.append(extracted_ref)
        
        return references
    
    def _deduplicate_references(self, references: List[ExtractedReference]) -> List[ExtractedReference]:
        """
        Remove duplicate references based on similarity
        """
        unique_references = []
        
        for ref in references:
            is_duplicate = False
            
            for existing_ref in unique_references:
                if self._are_references_similar(ref, existing_ref):
                    # Keep the one with higher confidence
                    if ref.confidence_score > existing_ref.confidence_score:
                        unique_references.remove(existing_ref)
                        unique_references.append(ref)
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                unique_references.append(ref)
        
        return unique_references
    
    def _are_references_similar(self, ref1: ExtractedReference, ref2: ExtractedReference) -> bool:
        """
        Check if two references are similar (likely duplicates)
        """
        # Compare titles if available
        if ref1.title and ref2.title:
            title_similarity = self._calculate_string_similarity(ref1.title, ref2.title)
            if title_similarity > 0.8:
                return True
        
        # Compare full text
        text_similarity = self._calculate_string_similarity(ref1.full_text, ref2.full_text)
        return text_similarity > 0.9
    
    def _calculate_string_similarity(self, str1: str, str2: str) -> float:
        """
        Calculate similarity between two strings (simple implementation)
        """
        if not str1 or not str2:
            return 0.0
        
        # Simple character-based similarity
        str1_lower = str1.lower()
        str2_lower = str2.lower()
        
        if str1_lower == str2_lower:
            return 1.0
        
        # Calculate Jaccard similarity on words
        words1 = set(str1_lower.split())
        words2 = set(str2_lower.split())
        
        intersection = words1.intersection(words2)
        union = words1.union(words2)
        
        if not union:
            return 0.0
        
        return len(intersection) / len(union)
    
    def _enhance_reference_quality(self, references: List[ExtractedReference]) -> List[ExtractedReference]:
        """
        Enhance the quality of extracted references
        """
        for ref in references:
            # Clean up text
            ref.full_text = self._clean_reference_text(ref.full_text)
            ref.title = self._clean_reference_text(ref.title)
            ref.venue = self._clean_reference_text(ref.venue)
            
            # Validate year
            if ref.year and (ref.year < 1900 or ref.year > 2030):
                ref.year = None
                ref.extraction_notes += "Invalid year detected. "
            
            # Update confidence score based on completeness
            ref.confidence_score = self._calculate_reference_completeness(ref)
        
        return references
    
    def _clean_reference_text(self, text: str) -> str:
        """
        Clean up reference text by removing extra whitespace, etc.
        """
        if not text:
            return ""
        
        # Remove extra whitespace
        text = re.sub(r'\s+', ' ', text)
        
        # Remove leading/trailing punctuation and whitespace
        text = text.strip(' .,;:')
        
        return text
    
    def _calculate_reference_completeness(self, ref: ExtractedReference) -> float:
        """
        Calculate a completeness score for the reference
        """
        score = 0.0
        total_fields = 0
        
        # Essential fields
        if ref.authors:
            score += 0.3
        total_fields += 1
        
        if ref.title:
            score += 0.3
        total_fields += 1
        
        if ref.year:
            score += 0.2
        total_fields += 1
        
        if ref.venue:
            score += 0.2
        total_fields += 1
        
        # Optional fields (smaller weight)
        optional_fields = [ref.doi, ref.url, ref.volume, ref.issue, ref.pages]
        filled_optional = sum(1 for field in optional_fields if field)
        score += (filled_optional / len(optional_fields)) * 0.2
        
        return min(1.0, score)

# Continue with the storage and main agent classes...


class ReferenceStorageManager:
    """
    Handles storage of extracted references in multiple formats (files and spreadsheets)
    """
    
    def __init__(self, config: Dict):
        self.config = config
        self.output_formats = ["json", "csv", "txt", "xlsx", "md"]
        
    def store_references(self, references: List[ExtractedReference], 
                        output_path: str, formats: List[str] = None) -> Dict[str, str]:
        """
        Store references in multiple formats
        
        Args:
            references: List of extracted references
            output_path: Base path for output files (without extension)
            formats: List of formats to generate (default: all)
        
        Returns:
            Dictionary mapping format to file path
        """
        if formats is None:
            formats = self.output_formats
        
        output_files = {}
        
        for format_type in formats:
            try:
                if format_type == "json":
                    file_path = self._save_as_json(references, f"{output_path}.json")
                elif format_type == "csv":
                    file_path = self._save_as_csv(references, f"{output_path}.csv")
                elif format_type == "txt":
                    file_path = self._save_as_text(references, f"{output_path}.txt")
                elif format_type == "xlsx":
                    file_path = self._save_as_excel(references, f"{output_path}.xlsx")
                elif format_type == "md":
                    file_path = self._save_as_markdown(references, f"{output_path}.md")
                else:
                    continue
                
                output_files[format_type] = file_path
                
            except Exception as e:
                print(f"Error saving as {format_type}: {str(e)}")
                output_files[format_type] = f"Error: {str(e)}"
        
        return output_files
    
    def _save_as_json(self, references: List[ExtractedReference], file_path: str) -> str:
        """Save references as JSON file"""
        references_data = []
        
        for ref in references:
            ref_dict = asdict(ref)
            # Convert any None values to empty strings for better JSON compatibility
            for key, value in ref_dict.items():
                if value is None:
                    ref_dict[key] = ""
            references_data.append(ref_dict)
        
        output_data = {
            "extraction_metadata": {
                "total_references": len(references),
                "extraction_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                "format_version": "1.0"
            },
            "references": references_data
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        
        return file_path
    
    def _save_as_csv(self, references: List[ExtractedReference], file_path: str) -> str:
        """Save references as CSV file"""
        if not references:
            return file_path
        
        # Define CSV columns
        columns = [
            'reference_number', 'full_text', 'authors', 'title', 'year', 'venue',
            'volume', 'issue', 'pages', 'doi', 'url', 'isbn', 'reference_type',
            'citation_style', 'confidence_score', 'extraction_notes'
        ]
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            
            for ref in references:
                row_data = {}
                for col in columns:
                    value = getattr(ref, col, "")
                    
                    # Handle list fields (like authors)
                    if isinstance(value, list):
                        value = "; ".join(str(item) for item in value)
                    elif value is None:
                        value = ""
                    
                    row_data[col] = str(value)
                
                writer.writerow(row_data)
        
        return file_path
    
    def _save_as_text(self, references: List[ExtractedReference], file_path: str) -> str:
        """Save references as formatted text file"""
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("EXTRACTED REFERENCES REPORT\\n")
            f.write("=" * 50 + "\\n\\n")
            f.write(f"Total References: {len(references)}\\n")
            f.write(f"Extraction Date: {time.strftime('%Y-%m-%d %H:%M:%S')}\\n\\n")
            
            for i, ref in enumerate(references, 1):
                f.write(f"REFERENCE {i}\\n")
                f.write("-" * 20 + "\\n")
                
                if ref.reference_number:
                    f.write(f"Number: {ref.reference_number}\\n")
                
                f.write(f"Full Text: {ref.full_text}\\n\\n")
                
                if ref.authors:
                    f.write(f"Authors: {'; '.join(ref.authors)}\\n")
                
                if ref.title:
                    f.write(f"Title: {ref.title}\\n")
                
                if ref.year:
                    f.write(f"Year: {ref.year}\\n")
                
                if ref.venue:
                    f.write(f"Venue: {ref.venue}\\n")
                
                if ref.volume:
                    f.write(f"Volume: {ref.volume}\\n")
                
                if ref.issue:
                    f.write(f"Issue: {ref.issue}\\n")
                
                if ref.pages:
                    f.write(f"Pages: {ref.pages}\\n")
                
                if ref.doi:
                    f.write(f"DOI: {ref.doi}\\n")
                
                if ref.url:
                    f.write(f"URL: {ref.url}\\n")
                
                if ref.isbn:
                    f.write(f"ISBN: {ref.isbn}\\n")
                
                if ref.reference_type:
                    f.write(f"Type: {ref.reference_type}\\n")
                
                if ref.citation_style:
                    f.write(f"Citation Style: {ref.citation_style}\\n")
                
                f.write(f"Confidence Score: {ref.confidence_score:.2f}\\n")
                
                if ref.extraction_notes:
                    f.write(f"Notes: {ref.extraction_notes}\\n")
                
                f.write("\\n" + "=" * 50 + "\\n\\n")
        
        return file_path
    
    def _save_as_excel(self, references: List[ExtractedReference], file_path: str) -> str:
        """Save references as Excel spreadsheet with formatting"""
        if not SPREADSHEET_LIBRARIES_AVAILABLE:
            raise ImportError("Spreadsheet libraries not available")
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted References"
        
        # Define headers
        headers = [
            'Ref #', 'Full Text', 'Authors', 'Title', 'Year', 'Venue',
            'Volume', 'Issue', 'Pages', 'DOI', 'URL', 'ISBN', 
            'Type', 'Style', 'Confidence', 'Notes'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, ref in enumerate(references, 2):
            ws.cell(row=row, column=1, value=ref.reference_number or "")
            ws.cell(row=row, column=2, value=ref.full_text)
            ws.cell(row=row, column=3, value="; ".join(ref.authors) if ref.authors else "")
            ws.cell(row=row, column=4, value=ref.title)
            ws.cell(row=row, column=5, value=ref.year)
            ws.cell(row=row, column=6, value=ref.venue)
            ws.cell(row=row, column=7, value=ref.volume)
            ws.cell(row=row, column=8, value=ref.issue)
            ws.cell(row=row, column=9, value=ref.pages)
            ws.cell(row=row, column=10, value=ref.doi)
            ws.cell(row=row, column=11, value=ref.url)
            ws.cell(row=row, column=12, value=ref.isbn)
            ws.cell(row=row, column=13, value=ref.reference_type)
            ws.cell(row=row, column=14, value=ref.citation_style)
            ws.cell(row=row, column=15, value=ref.confidence_score)
            ws.cell(row=row, column=16, value=ref.extraction_notes)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Summary data
        total_refs = len(references)
        avg_confidence = sum(ref.confidence_score for ref in references) / total_refs if total_refs > 0 else 0
        
        reference_types = {}
        citation_styles = {}
        
        for ref in references:
            ref_type = ref.reference_type or "unknown"
            reference_types[ref_type] = reference_types.get(ref_type, 0) + 1
            
            style = ref.citation_style or "unknown"
            citation_styles[style] = citation_styles.get(style, 0) + 1
        
        # Write summary
        summary_data = [
            ["Extraction Summary", ""],
            ["Total References", total_refs],
            ["Average Confidence", f"{avg_confidence:.2f}"],
            ["Extraction Date", time.strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Reference Types", "Count"],
        ]
        
        for ref_type, count in reference_types.items():
            summary_data.append([ref_type.title(), count])
        
        summary_data.extend([
            ["", ""],
            ["Citation Styles", "Count"],
        ])
        
        for style, count in citation_styles.items():
            summary_data.append([style.upper(), count])
        
        for row, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row, column=1, value=label)
            summary_ws.cell(row=row, column=2, value=value)
        
        # Style summary headers
        for row in [1, 6, len(reference_types) + 9]:
            if row <= len(summary_data):
                cell = summary_ws.cell(row=row, column=1)
                cell.font = Font(bold=True)
        
        # Save workbook
        wb.save(file_path)
        return file_path
    
    def _save_as_markdown(self, references: List[ExtractedReference], file_path: str) -> str:
        """Save references as Markdown file"""
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("# Extracted References Report\\n\\n")
            f.write(f"**Total References:** {len(references)}  \\n")
            f.write(f"**Extraction Date:** {time.strftime('%Y-%m-%d %H:%M:%S')}  \\n\\n")
            
            # Summary statistics
            if references:
                avg_confidence = sum(ref.confidence_score for ref in references) / len(references)
                f.write(f"**Average Confidence Score:** {avg_confidence:.2f}  \\n\\n")
                
                # Reference types summary
                ref_types = {}
                for ref in references:
                    ref_type = ref.reference_type or "unknown"
                    ref_types[ref_type] = ref_types.get(ref_type, 0) + 1
                
                f.write("## Reference Types Summary\\n\\n")
                for ref_type, count in sorted(ref_types.items()):
                    f.write(f"- **{ref_type.title()}:** {count}\\n")
                f.write("\\n")
            
            f.write("## Detailed References\\n\\n")
            
            for i, ref in enumerate(references, 1):
                f.write(f"### Reference {i}\\n\\n")
                
                if ref.reference_number:
                    f.write(f"**Reference Number:** {ref.reference_number}  \\n")
                
                f.write(f"**Full Text:** {ref.full_text}  \\n\\n")
                
                # Structured information table
                f.write("| Field | Value |\\n")
                f.write("|-------|-------|\\n")
                
                if ref.authors:
                    f.write(f"| Authors | {'; '.join(ref.authors)} |\\n")
                
                if ref.title:
                    f.write(f"| Title | {ref.title} |\\n")
                
                if ref.year:
                    f.write(f"| Year | {ref.year} |\\n")
                
                if ref.venue:
                    f.write(f"| Venue | {ref.venue} |\\n")
                
                if ref.volume:
                    f.write(f"| Volume | {ref.volume} |\\n")
                
                if ref.issue:
                    f.write(f"| Issue | {ref.issue} |\\n")
                
                if ref.pages:
                    f.write(f"| Pages | {ref.pages} |\\n")
                
                if ref.doi:
                    f.write(f"| DOI | {ref.doi} |\\n")
                
                if ref.url:
                    f.write(f"| URL | {ref.url} |\\n")
                
                if ref.isbn:
                    f.write(f"| ISBN | {ref.isbn} |\\n")
                
                if ref.reference_type:
                    f.write(f"| Type | {ref.reference_type} |\\n")
                
                if ref.citation_style:
                    f.write(f"| Citation Style | {ref.citation_style} |\\n")
                
                f.write(f"| Confidence Score | {ref.confidence_score:.2f} |\\n")
                
                if ref.extraction_notes:
                    f.write(f"| Notes | {ref.extraction_notes} |\\n")
                
                f.write("\\n---\\n\\n")
        
        return file_path

class PDFReferenceExtractionAgent:
    """
    Main agent class that orchestrates PDF text extraction, reference parsing, and storage
    """
    
    def __init__(self, config: Dict = None):
        if config is None:
            config = self._get_default_config()
        
        self.config = config
        
        # Initialize components
        self.text_extractor = PDFTextExtractor(config)
        self.reference_parser = ReferenceParser(config)
        self.storage_manager = ReferenceStorageManager(config)
        
        # Processing statistics
        self.processing_stats = {
            "total_pdfs_processed": 0,
            "total_references_extracted": 0,
            "average_confidence_score": 0.0,
            "processing_errors": []
        }
    
    def _get_default_config(self) -> Dict:
        """Get default configuration for the agent"""
        return {
            "extraction_methods": ["pdfplumber", "pypdf2", "ocr"],
            "output_formats": ["json", "csv", "xlsx", "txt", "md"],
            "min_reference_length": 20,
            "min_confidence_threshold": 0.3,
            "enable_ocr": True,
            "ocr_language": "eng",
            "max_file_size_mb": 50,
            "timeout_seconds": 300
        }
    
    def extract_references_from_pdf(self, pdf_path: str, output_path: str = None, 
                                   output_formats: List[str] = None) -> Dict[str, any]:
        """
        Main method to extract references from a PDF file
        
        Args:
            pdf_path: Path to the PDF file
            output_path: Base path for output files (without extension)
            output_formats: List of output formats to generate
        
        Returns:
            Dictionary containing extraction results and file paths
        """
        start_time = time.time()
        
        try:
            # Validate input
            if not os.path.exists(pdf_path):
                raise FileNotFoundError(f"PDF file not found: {pdf_path}")
            
            # Check file size
            file_size_mb = os.path.getsize(pdf_path) / (1024 * 1024)
            if file_size_mb > self.config["max_file_size_mb"]:
                raise ValueError(f"File too large: {file_size_mb:.1f}MB (max: {self.config['max_file_size_mb']}MB)")
            
            # Set default output path
            if output_path is None:
                pdf_name = Path(pdf_path).stem
                output_path = f"extracted_references_{pdf_name}"
            
            # Set default output formats
            if output_formats is None:
                output_formats = self.config["output_formats"]
            
            print(f"Processing PDF: {pdf_path}")
            print(f"File size: {file_size_mb:.1f}MB")
            
            # Step 1: Extract text from PDF
            print("Step 1: Extracting text from PDF...")
            text_extraction_result = self.text_extractor.extract_text_from_pdf(pdf_path)
            
            if not text_extraction_result["text"]:
                raise ValueError("No text could be extracted from the PDF")
            
            print(f"Text extraction completed using {text_extraction_result['method_used']}")
            print(f"Extraction quality: {text_extraction_result['extraction_quality']:.2f}")
            print(f"Total text length: {len(text_extraction_result['text'])} characters")
            
            # Step 2: Parse references from extracted text
            print("\\nStep 2: Parsing references from text...")
            references = self.reference_parser.extract_references_from_text(text_extraction_result["text"])
            
            # Filter references by confidence threshold
            filtered_references = [
                ref for ref in references 
                if ref.confidence_score >= self.config["min_confidence_threshold"]
            ]
            
            print(f"Found {len(references)} potential references")
            print(f"Filtered to {len(filtered_references)} references above confidence threshold")
            
            if filtered_references:
                avg_confidence = sum(ref.confidence_score for ref in filtered_references) / len(filtered_references)
                print(f"Average confidence score: {avg_confidence:.2f}")
            
            # Step 3: Store references in requested formats
            print("\\nStep 3: Storing references in output formats...")
            output_files = self.storage_manager.store_references(
                filtered_references, output_path, output_formats
            )
            
            # Calculate processing time
            processing_time = time.time() - start_time
            
            # Update statistics
            self.processing_stats["total_pdfs_processed"] += 1
            self.processing_stats["total_references_extracted"] += len(filtered_references)
            
            if filtered_references:
                current_avg = self.processing_stats["average_confidence_score"]
                total_processed = self.processing_stats["total_pdfs_processed"]
                new_avg = sum(ref.confidence_score for ref in filtered_references) / len(filtered_references)
                
                # Update running average
                self.processing_stats["average_confidence_score"] = (
                    (current_avg * (total_processed - 1) + new_avg) / total_processed
                )
            
            # Prepare result
            result = {
                "status": "success",
                "pdf_path": pdf_path,
                "processing_time": processing_time,
                "text_extraction": {
                    "method_used": text_extraction_result["method_used"],
                    "quality_score": text_extraction_result["extraction_quality"],
                    "text_length": len(text_extraction_result["text"])
                },
                "reference_extraction": {
                    "total_found": len(references),
                    "filtered_count": len(filtered_references),
                    "average_confidence": avg_confidence if filtered_references else 0.0,
                    "confidence_threshold": self.config["min_confidence_threshold"]
                },
                "references": filtered_references,
                "output_files": output_files,
                "file_size_mb": file_size_mb
            }
            
            print(f"\\nProcessing completed successfully in {processing_time:.2f} seconds")
            print(f"Output files generated: {list(output_files.keys())}")
            
            return result
            
        except Exception as e:
            error_msg = str(e)
            self.processing_stats["processing_errors"].append({
                "pdf_path": pdf_path,
                "error": error_msg,
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S")
            })
            
            print(f"Error processing PDF: {error_msg}")
            
            return {
                "status": "error",
                "pdf_path": pdf_path,
                "error": error_msg,
                "processing_time": time.time() - start_time
            }
    
    def batch_extract_references(self, pdf_paths: List[str], output_dir: str = "batch_extraction") -> Dict[str, any]:
        """
        Extract references from multiple PDF files in batch
        
        Args:
            pdf_paths: List of paths to PDF files
            output_dir: Directory to store all output files
        
        Returns:
            Dictionary containing batch processing results
        """
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        batch_results = {}
        successful_extractions = 0
        failed_extractions = 0
        
        print(f"Starting batch extraction of {len(pdf_paths)} PDF files...")
        print(f"Output directory: {output_dir}")
        
        for i, pdf_path in enumerate(pdf_paths, 1):
            print(f"\\n{'='*60}")
            print(f"Processing file {i}/{len(pdf_paths)}: {os.path.basename(pdf_path)}")
            print(f"{'='*60}")
            
            # Generate output path for this PDF
            pdf_name = Path(pdf_path).stem
            output_path = os.path.join(output_dir, f"references_{pdf_name}")
            
            # Extract references
            result = self.extract_references_from_pdf(pdf_path, output_path)
            batch_results[pdf_path] = result
            
            if result["status"] == "success":
                successful_extractions += 1
            else:
                failed_extractions += 1
        
        # Generate batch summary
        total_references = sum(
            len(result.get("references", [])) 
            for result in batch_results.values() 
            if result["status"] == "success"
        )
        
        batch_summary = {
            "total_files": len(pdf_paths),
            "successful_extractions": successful_extractions,
            "failed_extractions": failed_extractions,
            "total_references_extracted": total_references,
            "output_directory": output_dir,
            "processing_timestamp": time.strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # Save batch summary
        summary_path = os.path.join(output_dir, "batch_summary.json")
        with open(summary_path, 'w', encoding='utf-8') as f:
            json.dump({
                "batch_summary": batch_summary,
                "individual_results": {
                    path: {
                        "status": result["status"],
                        "references_count": len(result.get("references", [])),
                        "processing_time": result.get("processing_time", 0),
                        "error": result.get("error")
                    }
                    for path, result in batch_results.items()
                }
            }, f, indent=2)
        
        print(f"\\n{'='*60}")
        print("BATCH PROCESSING COMPLETE")
        print(f"{'='*60}")
        print(f"Total files processed: {len(pdf_paths)}")
        print(f"Successful extractions: {successful_extractions}")
        print(f"Failed extractions: {failed_extractions}")
        print(f"Total references extracted: {total_references}")
        print(f"Batch summary saved to: {summary_path}")
        
        return {
            "batch_summary": batch_summary,
            "individual_results": batch_results,
            "summary_file": summary_path
        }
    
    def get_processing_statistics(self) -> Dict[str, any]:
        """Get current processing statistics"""
        return self.processing_stats.copy()
    
    def reset_statistics(self):
        """Reset processing statistics"""
        self.processing_stats = {
            "total_pdfs_processed": 0,
            "total_references_extracted": 0,
            "average_confidence_score": 0.0,
            "processing_errors": []
        }

# Integration with ARAS system
def integrate_with_aras(pdf_extraction_result: Dict[str, any], aras_system) -> Dict[str, any]:
    """
    Integrate PDF extraction results with the Academic Research Agent System
    for enhanced verification and analysis
    """
    if pdf_extraction_result["status"] != "success":
        return pdf_extraction_result
    
    references = pdf_extraction_result["references"]
    enhanced_references = []
    
    print("\\nIntegrating with ARAS for enhanced verification...")
    
    for ref in references:
        # Convert ExtractedReference to citation string for ARAS validation
        citation_text = ref.full_text
        
        # Use ARAS to validate and enhance the reference
        try:
            validation_result = aras_system.validate_citations([citation_text])
            
            # Update reference with ARAS insights
            if validation_result["status"] == "success":
                ref.extraction_notes += " ARAS validation completed."
                
                # Add any corrections or suggestions from ARAS
                if validation_result.get("corrections"):
                    ref.extraction_notes += f" ARAS suggestions: {validation_result['corrections']}"
            
        except Exception as e:
            ref.extraction_notes += f" ARAS integration error: {str(e)}"
        
        enhanced_references.append(ref)
    
    # Update the result with enhanced references
    pdf_extraction_result["references"] = enhanced_references
    pdf_extraction_result["aras_integration"] = True
    
    return pdf_extraction_result

# Demonstration and testing functions
def demonstrate_pdf_extraction():
    """
    Demonstrate the PDF Reference Extraction Agent capabilities
    """
    print("=" * 60)
    print("PDF REFERENCE EXTRACTION AGENT DEMONSTRATION")
    print("=" * 60)
    
    # Initialize the agent
    config = {
        "min_confidence_threshold": 0.3,
        "output_formats": ["json", "csv", "xlsx", "txt", "md"]
    }
    
    agent = PDFReferenceExtractionAgent(config)
    print("✓ PDF Reference Extraction Agent initialized")
    
    # Since we don't have actual PDFs to test with, we'll simulate the process
    print("\\n" + "=" * 60)
    print("SIMULATED EXTRACTION PROCESS")
    print("=" * 60)
    
    # Simulate extracted references
    sample_references = [
        ExtractedReference(
            reference_number=1,
            full_text="Smith, J. A. (2023). Machine learning applications in academic research. Journal of AI Research, 15(3), 45-62.",
            authors=["Smith, J. A."],
            title="Machine learning applications in academic research",
            year=2023,
            venue="Journal of AI Research",
            volume="15",
            issue="3",
            pages="45-62",
            reference_type="journal",
            citation_style="apa",
            confidence_score=0.95
        ),
        ExtractedReference(
            reference_number=2,
            full_text="Johnson, M., & Brown, K. (2022). Data Science Fundamentals. Academic Press.",
            authors=["Johnson, M.", "Brown, K."],
            title="Data Science Fundamentals",
            year=2022,
            venue="Academic Press",
            reference_type="book",
            citation_style="apa",
            confidence_score=0.88
        ),
        ExtractedReference(
            reference_number=3,
            full_text="Wilson, P. et al. (2024). 'Automated citation analysis', Proceedings of the International Conference on Digital Libraries, pp. 123-135.",
            authors=["Wilson, P.", "et al."],
            title="Automated citation analysis",
            year=2024,
            venue="Proceedings of the International Conference on Digital Libraries",
            pages="123-135",
            reference_type="conference",
            citation_style="apa",
            confidence_score=0.82
        )
    ]
    
    print(f"Simulated extraction of {len(sample_references)} references:")
    
    for ref in sample_references:
        print(f"\\n[{ref.reference_number}] {ref.title}")
        print(f"    Authors: {', '.join(ref.authors)}")
        print(f"    Year: {ref.year}")
        print(f"    Type: {ref.reference_type}")
        print(f"    Confidence: {ref.confidence_score:.2f}")
    
    # Demonstrate storage capabilities
    print("\\n" + "=" * 60)
    print("STORAGE DEMONSTRATION")
    print("=" * 60)
    
    storage_manager = ReferenceStorageManager(config)
    
    # Store in different formats
    output_files = storage_manager.store_references(
        sample_references, 
        "demo_references", 
        ["json", "csv", "txt", "md"]
    )
    
    print("References stored in multiple formats:")
    for format_type, file_path in output_files.items():
        if not file_path.startswith("Error"):
            print(f"✓ {format_type.upper()}: {file_path}")
        else:
            print(f"✗ {format_type.upper()}: {file_path}")
    
    # Display statistics
    print("\\n" + "=" * 60)
    print("EXTRACTION STATISTICS")
    print("=" * 60)
    
    total_refs = len(sample_references)
    avg_confidence = sum(ref.confidence_score for ref in sample_references) / total_refs
    
    ref_types = {}
    for ref in sample_references:
        ref_type = ref.reference_type
        ref_types[ref_type] = ref_types.get(ref_type, 0) + 1
    
    print(f"Total References: {total_refs}")
    print(f"Average Confidence: {avg_confidence:.2f}")
    print(f"Reference Types: {dict(ref_types)}")
    
    print("\\n" + "=" * 60)
    print("DEMONSTRATION COMPLETE")
    print("=" * 60)
    
    return {
        "sample_references": sample_references,
        "output_files": output_files,
        "statistics": {
            "total_references": total_refs,
            "average_confidence": avg_confidence,
            "reference_types": ref_types
        }
    }

if __name__ == "__main__":
    # Run demonstration
    demo_result = demonstrate_pdf_extraction()
    
    print("\\nThe PDF Reference Extraction Agent is ready for use!")
    print("Key capabilities:")
    print("• Extract text from PDFs using multiple methods")
    print("• Parse academic references with high accuracy")
    print("• Store results in JSON, CSV, Excel, Text, and Markdown formats")
    print("• Batch process multiple PDF files")
    print("• Integration with ARAS for enhanced verification")

