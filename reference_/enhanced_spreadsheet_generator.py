"""
Enhanced Reference Processing and Spreadsheet Generation
Creating comprehensive Excel spreadsheets from the extracted references
"""

import json
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference

def create_enhanced_spreadsheets():
    """Create comprehensive Excel spreadsheets from the extracted references"""
    
    print("=" * 70)
    print("ENHANCED REFERENCE PROCESSING AND SPREADSHEET GENERATION")
    print("=" * 70)
    
    # Load the extracted references
    print("\\n1. Loading extracted reference data...")
    
    try:
        with open('/home/ubuntu/research_report_reference_analysis.json', 'r') as f:
            analysis_data = json.load(f)
        
        references = analysis_data['references']
        statistics = analysis_data['statistics']
        
        print(f"✓ Loaded {len(references)} references from analysis file")
        
    except Exception as e:
        print(f"✗ Error loading reference data: {e}")
        return None
    
    # Create comprehensive DataFrame
    print("\\n2. Creating comprehensive reference DataFrame...")
    
    df = create_reference_dataframe(references)
    print(f"✓ Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
    
    # Generate enhanced Excel workbook
    print("\\n3. Generating enhanced Excel workbook...")
    
    excel_file = create_enhanced_excel_workbook(df, statistics, analysis_data)
    print(f"✓ Created enhanced Excel workbook: {os.path.basename(excel_file)}")
    
    # Generate summary CSV
    print("\\n4. Generating summary CSV files...")
    
    csv_files = create_summary_csv_files(df, statistics)
    print(f"✓ Created {len(csv_files)} summary CSV files")
    
    # Generate analysis report
    print("\\n5. Generating detailed analysis report...")
    
    report_file = create_analysis_report(df, statistics, analysis_data)
    print(f"✓ Created analysis report: {os.path.basename(report_file)}")
    
    return {
        "excel_file": excel_file,
        "csv_files": csv_files,
        "report_file": report_file,
        "dataframe": df,
        "statistics": statistics
    }

def create_reference_dataframe(references):
    """Create a comprehensive pandas DataFrame from the references"""
    
    data = []
    
    for ref in references:
        # Handle authors list
        authors_str = ", ".join(ref.get('authors', [])) if ref.get('authors') else ""
        first_author = ref.get('authors', [''])[0] if ref.get('authors') else ""
        author_count = len(ref.get('authors', []))
        
        # Extract domain from URL if available
        url = ref.get('url', '')
        domain = ""
        if url:
            try:
                from urllib.parse import urlparse
                domain = urlparse(url).netloc
            except:
                domain = ""
        
        # Determine citation style
        full_text = ref.get('full_text', '')
        citation_style = determine_citation_style(full_text)
        
        row = {
            'Reference_Number': ref.get('reference_number', ''),
            'Title': ref.get('title', ''),
            'Authors': authors_str,
            'First_Author': first_author,
            'Author_Count': author_count,
            'Year': ref.get('year', ''),
            'Venue': ref.get('venue', ''),
            'Volume': ref.get('volume', ''),
            'Issue': ref.get('issue', ''),
            'Pages': ref.get('pages', ''),
            'DOI': ref.get('doi', ''),
            'URL': url,
            'Domain': domain,
            'ISBN': ref.get('isbn', ''),
            'Reference_Type': ref.get('reference_type', ''),
            'Citation_Style': citation_style,
            'Confidence_Score': ref.get('confidence_score', 0),
            'Full_Text': full_text,
            'Text_Length': len(full_text),
            'Has_DOI': bool(ref.get('doi')),
            'Has_URL': bool(url),
            'Has_ISBN': bool(ref.get('isbn')),
            'Decade': (ref.get('year') // 10 * 10) if ref.get('year') else None
        }
        
        data.append(row)
    
    return pd.DataFrame(data)

def determine_citation_style(full_text):
    """Determine the likely citation style based on the text format"""
    
    if not full_text:
        return "Unknown"
    
    # Simple heuristics for citation style detection
    if "doi:" in full_text.lower() or "DOI:" in full_text:
        return "APA"
    elif full_text.count("(") > 1 and full_text.count(")") > 1:
        return "APA"
    elif '"' in full_text and full_text.count('"') >= 2:
        return "MLA"
    elif "vol." in full_text.lower() or "no." in full_text.lower():
        return "Chicago"
    elif "pp." in full_text:
        return "MLA"
    else:
        return "Mixed"

def create_enhanced_excel_workbook(df, statistics, analysis_data):
    """Create an enhanced Excel workbook with multiple sheets and formatting"""
    
    filename = "/home/ubuntu/research_report_references_enhanced.xlsx"
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # 1. Summary Sheet
    create_summary_sheet(wb, df, statistics, analysis_data)
    
    # 2. All References Sheet
    create_references_sheet(wb, df)
    
    # 3. By Type Sheet
    create_by_type_sheet(wb, df)
    
    # 4. By Year Sheet
    create_by_year_sheet(wb, df)
    
    # 5. Quality Analysis Sheet
    create_quality_analysis_sheet(wb, df)
    
    # 6. Author Analysis Sheet
    create_author_analysis_sheet(wb, df)
    
    # Save workbook
    wb.save(filename)
    
    return filename

def create_summary_sheet(wb, df, statistics, analysis_data):
    """Create a summary sheet with key statistics and charts"""
    
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = "Research Report Reference Analysis Summary"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Basic statistics
    row = 3
    ws[f'A{row}'] = "Extraction Statistics"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    stats_data = [
        ("Total References", statistics['total_references']),
        ("Average Confidence Score", f"{statistics['average_confidence']:.2f}"),
        ("Year Range", statistics['year_range']),
        ("Unique Authors", statistics['unique_authors']),
        ("Unique Venues", statistics['unique_venues']),
        ("Extraction Date", analysis_data['extraction_metadata']['extraction_timestamp'])
    ]
    
    for i, (label, value) in enumerate(stats_data, row + 1):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Reference type breakdown
    row += len(stats_data) + 3
    ws[f'A{row}'] = "Reference Types"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    for i, (ref_type, count) in enumerate(statistics['reference_types'].items(), row + 1):
        ws[f'A{i}'] = ref_type.title()
        ws[f'B{i}'] = count
        ws[f'C{i}'] = f"{count/statistics['total_references']*100:.1f}%"
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_references_sheet(wb, df):
    """Create a sheet with all references in a formatted table"""
    
    ws = wb.create_sheet("All References")
    
    # Add title
    ws['A1'] = "Complete Reference List"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Select key columns for display
    display_columns = [
        'Reference_Number', 'Title', 'Authors', 'Year', 'Venue', 
        'Reference_Type', 'Confidence_Score', 'Has_DOI', 'Has_URL'
    ]
    
    display_df = df[display_columns].copy()
    
    # Add data to worksheet
    for r in dataframe_to_rows(display_df, index=False, header=True):
        ws.append(r)
    
    # Format header row
    for cell in ws[3]:  # Row 3 contains headers
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_by_type_sheet(wb, df):
    """Create a sheet analyzing references by type"""
    
    ws = wb.create_sheet("By Type")
    
    # Title
    ws['A1'] = "References by Type Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Type summary
    type_summary = df.groupby('Reference_Type').agg({
        'Reference_Number': 'count',
        'Confidence_Score': 'mean',
        'Year': ['min', 'max'],
        'Has_DOI': 'sum',
        'Has_URL': 'sum'
    }).round(2)
    
    # Flatten column names
    type_summary.columns = ['Count', 'Avg_Confidence', 'Min_Year', 'Max_Year', 'DOI_Count', 'URL_Count']
    type_summary = type_summary.reset_index()
    
    # Add to worksheet
    row = 3
    for r in dataframe_to_rows(type_summary, index=False, header=True):
        ws.append(r)
    
    # Format header
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

def create_by_year_sheet(wb, df):
    """Create a sheet analyzing references by year"""
    
    ws = wb.create_sheet("By Year")
    
    # Title
    ws['A1'] = "References by Year Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Year summary
    year_summary = df[df['Year'].notna()].groupby('Year').agg({
        'Reference_Number': 'count',
        'Confidence_Score': 'mean',
        'Reference_Type': lambda x: x.mode().iloc[0] if not x.mode().empty else 'Unknown'
    }).round(2)
    
    year_summary.columns = ['Count', 'Avg_Confidence', 'Most_Common_Type']
    year_summary = year_summary.reset_index()
    year_summary = year_summary.sort_values('Year')
    
    # Add to worksheet
    row = 3
    for r in dataframe_to_rows(year_summary, index=False, header=True):
        ws.append(r)
    
    # Format header
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

def create_quality_analysis_sheet(wb, df):
    """Create a sheet analyzing reference quality"""
    
    ws = wb.create_sheet("Quality Analysis")
    
    # Title
    ws['A1'] = "Reference Quality Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Quality metrics
    row = 3
    quality_metrics = [
        ("High Confidence (>0.8)", len(df[df['Confidence_Score'] > 0.8])),
        ("Medium Confidence (0.5-0.8)", len(df[(df['Confidence_Score'] >= 0.5) & (df['Confidence_Score'] <= 0.8)])),
        ("Low Confidence (<0.5)", len(df[df['Confidence_Score'] < 0.5])),
        ("", ""),
        ("References with DOI", len(df[df['Has_DOI'] == True])),
        ("References with URL", len(df[df['Has_URL'] == True])),
        ("References with ISBN", len(df[df['Has_ISBN'] == True])),
        ("", ""),
        ("Complete Author Info", len(df[df['Author_Count'] > 0])),
        ("Complete Year Info", len(df[df['Year'].notna()])),
        ("Complete Venue Info", len(df[df['Venue'].notna() & (df['Venue'] != '')]))
    ]
    
    for i, (metric, value) in enumerate(quality_metrics, row):
        ws[f'A{i}'] = metric
        ws[f'B{i}'] = value
        if metric:  # Don't bold empty rows
            ws[f'A{i}'].font = Font(bold=True)

def create_author_analysis_sheet(wb, df):
    """Create a sheet analyzing author information"""
    
    ws = wb.create_sheet("Author Analysis")
    
    # Title
    ws['A1'] = "Author Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Author statistics
    row = 3
    author_stats = [
        ("Total Unique Authors", df['Author_Count'].sum()),
        ("References with Authors", len(df[df['Author_Count'] > 0])),
        ("Single Author Papers", len(df[df['Author_Count'] == 1])),
        ("Multi-Author Papers", len(df[df['Author_Count'] > 1])),
        ("Average Authors per Paper", df['Author_Count'].mean()),
        ("Max Authors in Single Paper", df['Author_Count'].max())
    ]
    
    for i, (stat, value) in enumerate(author_stats, row):
        ws[f'A{i}'] = stat
        ws[f'B{i}'] = f"{value:.2f}" if isinstance(value, float) else value
        ws[f'A{i}'].font = Font(bold=True)

def create_summary_csv_files(df, statistics):
    """Create summary CSV files for different analyses"""
    
    csv_files = {}
    
    # 1. Basic reference summary
    summary_df = df[['Reference_Number', 'Title', 'Authors', 'Year', 'Venue', 'Reference_Type', 'Confidence_Score']].copy()
    summary_file = "/home/ubuntu/research_report_references_summary.csv"
    summary_df.to_csv(summary_file, index=False)
    csv_files['summary'] = summary_file
    
    # 2. Type analysis
    type_analysis = df.groupby('Reference_Type').agg({
        'Reference_Number': 'count',
        'Confidence_Score': 'mean',
        'Year': ['min', 'max']
    }).round(2)
    type_file = "/home/ubuntu/research_report_references_by_type.csv"
    type_analysis.to_csv(type_file)
    csv_files['by_type'] = type_file
    
    # 3. Year analysis
    year_analysis = df[df['Year'].notna()].groupby('Year').size().reset_index(name='Count')
    year_file = "/home/ubuntu/research_report_references_by_year.csv"
    year_analysis.to_csv(year_file, index=False)
    csv_files['by_year'] = year_file
    
    # 4. Quality metrics
    quality_df = df[['Reference_Number', 'Title', 'Confidence_Score', 'Has_DOI', 'Has_URL', 'Has_ISBN', 'Author_Count']].copy()
    quality_file = "/home/ubuntu/research_report_references_quality.csv"
    quality_df.to_csv(quality_file, index=False)
    csv_files['quality'] = quality_file
    
    return csv_files

def create_analysis_report(df, statistics, analysis_data):
    """Create a detailed text analysis report"""
    
    report_file = "/home/ubuntu/research_report_reference_analysis_report.txt"
    
    with open(report_file, 'w') as f:
        f.write("RESEARCH REPORT REFERENCE ANALYSIS\\n")
        f.write("=" * 50 + "\\n\\n")
        
        # Basic statistics
        f.write("EXTRACTION SUMMARY\\n")
        f.write("-" * 20 + "\\n")
        f.write(f"Total References Extracted: {statistics['total_references']}\\n")
        f.write(f"Average Confidence Score: {statistics['average_confidence']:.2f}\\n")
        f.write(f"Year Range: {statistics['year_range']}\\n")
        f.write(f"Unique Authors: {statistics['unique_authors']}\\n")
        f.write(f"Unique Venues: {statistics['unique_venues']}\\n\\n")
        
        # Reference types
        f.write("REFERENCE TYPES\\n")
        f.write("-" * 15 + "\\n")
        for ref_type, count in statistics['reference_types'].items():
            percentage = count / statistics['total_references'] * 100
            f.write(f"{ref_type.title()}: {count} ({percentage:.1f}%)\\n")
        f.write("\\n")
        
        # Quality analysis
        f.write("QUALITY ANALYSIS\\n")
        f.write("-" * 15 + "\\n")
        high_conf = len(df[df['Confidence_Score'] > 0.8])
        med_conf = len(df[(df['Confidence_Score'] >= 0.5) & (df['Confidence_Score'] <= 0.8)])
        low_conf = len(df[df['Confidence_Score'] < 0.5])
        
        f.write(f"High Confidence (>0.8): {high_conf} ({high_conf/len(df)*100:.1f}%)\\n")
        f.write(f"Medium Confidence (0.5-0.8): {med_conf} ({med_conf/len(df)*100:.1f}%)\\n")
        f.write(f"Low Confidence (<0.5): {low_conf} ({low_conf/len(df)*100:.1f}%)\\n\\n")
        
        # Metadata completeness
        f.write("METADATA COMPLETENESS\\n")
        f.write("-" * 20 + "\\n")
        f.write(f"References with DOI: {df['Has_DOI'].sum()} ({df['Has_DOI'].sum()/len(df)*100:.1f}%)\\n")
        f.write(f"References with URL: {df['Has_URL'].sum()} ({df['Has_URL'].sum()/len(df)*100:.1f}%)\\n")
        f.write(f"References with ISBN: {df['Has_ISBN'].sum()} ({df['Has_ISBN'].sum()/len(df)*100:.1f}%)\\n")
        f.write(f"References with Authors: {(df['Author_Count'] > 0).sum()} ({(df['Author_Count'] > 0).sum()/len(df)*100:.1f}%)\\n")
        f.write(f"References with Year: {df['Year'].notna().sum()} ({df['Year'].notna().sum()/len(df)*100:.1f}%)\\n\\n")
        
        # Top venues
        f.write("TOP VENUES\\n")
        f.write("-" * 10 + "\\n")
        venue_counts = df[df['Venue'].notna() & (df['Venue'] != '')]['Venue'].value_counts().head(10)
        for venue, count in venue_counts.items():
            f.write(f"{venue}: {count}\\n")
        f.write("\\n")
        
        # Year distribution
        f.write("YEAR DISTRIBUTION\\n")
        f.write("-" * 15 + "\\n")
        year_counts = df[df['Year'].notna()]['Year'].value_counts().sort_index()
        for year, count in year_counts.items():
            f.write(f"{int(year)}: {count}\\n")
    
    return report_file

if __name__ == "__main__":
    result = create_enhanced_spreadsheets()
    
    if result:
        print("\\n" + "=" * 70)
        print("ENHANCED SPREADSHEET GENERATION COMPLETED")
        print("=" * 70)
        
        print(f"\\n Excel Workbook: {os.path.basename(result['excel_file'])}")
        print(f" Analysis Report: {os.path.basename(result['report_file'])}")
        
        print("\\n CSV Files Generated:")
        for csv_type, csv_file in result['csv_files'].items():
            print(f"  - {csv_type.title()}: {os.path.basename(csv_file)}")
        
        print("\\n Key Statistics:")
        stats = result['statistics']
        print(f"  - Total References: {stats['total_references']}")
        print(f"  - Average Confidence: {stats['average_confidence']:.2f}")
        print(f"  - Reference Types: {len(stats['reference_types'])}")
        print(f"  - Year Range: {stats['year_range']}")
        
        print("\\n All files generated successfully!")
        
    else:
        print("\\n Spreadsheet generation failed.")

